# 该脚本是把所有的非空单元格的数据放到第一列
import openpyxl
import time

# 记录开始时间
start_time = time.time()

# 读取 Excel 文件
workbook = openpyxl.load_workbook("AU.xlsx")
sheet = workbook.active

# 创建一个新的 Excel 文件
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 遍历每一行
for row in sheet.iter_rows(values_only=True):
    # 遍历每个单元格，并将其值添加到列表中（如果不是空值的话）
    for cell_value in row:
        if cell_value is not None:
            new_sheet.append([cell_value])

# 保存结果到新的 Excel 文件
new_workbook.save("output_excel_file.xlsx")

# 记录结束时间
end_time = time.time()

# 计算运行时间
running_time = end_time - start_time
print("代码运行时间：", running_time, "秒")


