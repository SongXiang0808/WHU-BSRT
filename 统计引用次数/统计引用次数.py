import openpyxl

# 读取 Excel 文件
workbook = openpyxl.load_workbook("CT.xlsx")
sheet = workbook.active

# 创建一个空字典，用于存储每个期刊的引用次数
journal_citations = {}

# 遍历每一行
for row in sheet.iter_rows(values_only=True):
    journal_name = row[0]
    citations = row[1]

    # 如果期刊名字已经在字典中，则累加引用次数；否则将其添加到字典中
    if journal_name in journal_citations:
        journal_citations[journal_name] += citations
    else:
        journal_citations[journal_name] = citations

# 创建一个新的 Excel 文件
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 将统计结果写入到新的 Excel 文件中
for i, (journal, citations) in enumerate(journal_citations.items(), start=1):
    new_sheet.cell(row=i, column=1, value=journal)
    new_sheet.cell(row=i, column=2, value=citations)

# 保存结果到新的 Excel 文件
new_workbook.save("journal_citations.xlsx")
