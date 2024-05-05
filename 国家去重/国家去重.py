import openpyxl
import time

# 记录开始时间
start_time = time.time()

# 读取 Excel 文件
workbook = openpyxl.load_workbook("作者国家表.xlsx")
sheet = workbook.active

# 创建一个新的 Excel 文件
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 遍历每一行
for row in sheet.iter_rows(values_only=True):
    # 创建一个空的集合，用于存放每一行中已经出现过的值
    unique_values = set()

    # 创建一个空的列表，用于存放当前行中去重后的值
    unique_row = []

    # 遍历当前行中的每个单元格
    for cell in row:
        # 如果单元格的值在集合中不存在，则将其添加到集合和列表中
        if cell not in unique_values:
            unique_values.add(cell)
            unique_row.append(cell)

    # 将去重后的行写入到新的 Excel 文件中
    new_sheet.append(unique_row)

# 保存结果到新的 Excel 文件
new_workbook.save("去重作者国家表.xlsx")

# 记录结束时间
end_time = time.time()

# 计算运行时间
running_time = end_time - start_time
print("代码运行时间：", running_time, "秒")

